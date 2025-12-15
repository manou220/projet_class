"""
Blueprint pour la gestion des résultats de tests et téléchargements.
"""
import json
from datetime import datetime
from io import BytesIO
from flask import Blueprint, render_template, session, jsonify, send_file, current_app, request
from app.utils import get_history

bp = Blueprint('resultats', __name__)


@bp.route('/')
def index():
    """Affiche la page de résultats."""
    if 'last_result' in session:
        last_result = session['last_result']
        results_text = f"""
Test : {last_result['test_name'].upper()}
Fichier : {last_result['filename']}
Colonnes : {', '.join(last_result['columns_used'])}
Statistique : {last_result['stat_value']:.6f}
P-value : {last_result['p_value']:.6f}

Interprétation : {last_result['interpretation']}
"""
        
        return render_template('resultats.html',
                             test_name=last_result['test_name'],
                             column=', '.join(last_result['columns_used']),
                             results_text=results_text,
                             interpretation=last_result['interpretation'],
                             can_download=True)
    
    return render_template("resultats.html")


@bp.route("/download_last_result")
def download_last_result():
    """Télécharge le dernier résultat de test au format JSON."""
    if 'last_result' not in session:
        return "Aucun résultat récent à télécharger.", 404

    result_data = session['last_result']
    result_json = json.dumps(result_data, indent=4, ensure_ascii=False)
    
    buffer = BytesIO()
    buffer.write(result_json.encode('utf-8'))
    buffer.seek(0)
    
    filename = f"resultat_{result_data['test_name']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/json')


@bp.route("/download_history")
def download_history():
    """Télécharge l'historique complet des tests au format JSON, PDF ou Excel."""
    from app.utils import get_history
    history = get_history(session=session)
    
    if not history:
        return "Aucun historique à télécharger.", 404
    
    # Récupérer le format demandé (par défaut JSON)
    format_type = request.args.get('format', 'json').lower()
    
    if format_type == 'pdf':
        return download_history_pdf(history)
    elif format_type == 'excel':
        return download_history_excel(history)
    else:
        # Format JSON par défaut
        history_json = json.dumps(history, indent=4, ensure_ascii=False)
        
        buffer = BytesIO()
        buffer.write(history_json.encode('utf-8'))
        buffer.seek(0)
        
        return send_file(buffer, as_attachment=True,
                        download_name=f'historique_tests_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json',
                        mimetype='application/json')


def download_history_pdf(history):
    """Génère un PDF de l'historique des tests."""
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # Titre
        story.append(Paragraph("Historique des Tests Statistiques", title_style))
        story.append(Spacer(1, 0.2*inch))
        
        # Informations générales
        info_text = f"<b>Date de génération:</b> {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}<br/>"
        info_text += f"<b>Nombre total de tests:</b> {len(history)}"
        story.append(Paragraph(info_text, styles['Normal']))
        story.append(Spacer(1, 0.3*inch))
        
        # Tableau des résultats
        data = [['Test', 'Fichier', 'Colonnes', 'P-Value', 'Statistique', 'Date']]
        
        for test in history:
            columns_str = ', '.join(test.get('columns_used', []))[:30] + '...' if len(', '.join(test.get('columns_used', []))) > 30 else ', '.join(test.get('columns_used', []))
            data.append([
                test.get('test_name', 'N/A'),
                test.get('filename', 'N/A')[:20],
                columns_str,
                f"{test.get('p_value', 0):.4f}",
                f"{test.get('stat_value', 0):.4f}",
                test.get('timestamp', 'N/A')[:10]
            ])
        
        table = Table(data, colWidths=[1.2*inch, 1.2*inch, 1.5*inch, 0.8*inch, 1*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        story.append(table)
        
        # Générer le PDF
        doc.build(story)
        buffer.seek(0)
        
        filename = f'historique_tests_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/pdf')
        
    except ImportError:
        current_app.logger.error("reportlab n'est pas installé. Installez-le avec: pip install reportlab")
        return jsonify({'error': 'Export PDF non disponible. Installez reportlab.'}), 500
    except Exception as e:
        current_app.logger.exception(f"Erreur génération PDF: {e}")
        return jsonify({'error': f'Erreur lors de la génération du PDF: {str(e)}'}), 500


def download_history_excel(history):
    """Génère un fichier Excel de l'historique des tests."""
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        buffer = BytesIO()
        
        # Créer un DataFrame
        data = []
        for test in history:
            data.append({
                'Test': test.get('test_name', 'N/A'),
                'Fichier': test.get('filename', 'N/A'),
                'Colonnes': ', '.join(test.get('columns_used', [])),
                'P-Value': test.get('p_value', 0),
                'Statistique': test.get('stat_value', 0),
                'Interprétation': test.get('interpretation', 'N/A'),
                'Date': test.get('timestamp', 'N/A')
            })
        
        df = pd.DataFrame(data)
        
        # Écrire dans Excel avec openpyxl pour le formatage
        wb = Workbook()
        ws = wb.active
        ws.title = "Historique Tests"
        
        # En-têtes
        headers = ['Test', 'Fichier', 'Colonnes', 'P-Value', 'Statistique', 'Interprétation', 'Date']
        ws.append(headers)
        
        # Style des en-têtes
        header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Données
        for idx, row in df.iterrows():
            ws.append([
                row['Test'],
                row['Fichier'],
                row['Colonnes'],
                row['P-Value'],
                row['Statistique'],
                row['Interprétation'],
                row['Date']
            ])
            
            # Appliquer les bordures
            for cell in ws[idx + 2]:
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Ajuster la largeur des colonnes
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 40
        ws.column_dimensions['G'].width = 20
        
        # Sauvegarder dans le buffer
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f'historique_tests_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return send_file(buffer, as_attachment=True, download_name=filename, 
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
    except ImportError:
        current_app.logger.error("openpyxl n'est pas installé. Installez-le avec: pip install openpyxl")
        return jsonify({'error': 'Export Excel non disponible. Installez openpyxl.'}), 500
    except Exception as e:
        current_app.logger.exception(f"Erreur génération Excel: {e}")
        return jsonify({'error': f'Erreur lors de la génération du Excel: {str(e)}'}), 500
